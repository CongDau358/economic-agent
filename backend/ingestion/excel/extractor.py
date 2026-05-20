from __future__ import annotations

import csv
from dataclasses import dataclass, field
from pathlib import Path
from typing import List, Sequence


@dataclass
class ExtractedTable:
    sheet_name: str
    headers: List[str]
    rows: List[List[str]]
    row_start: int
    row_end: int
    col_count: int


@dataclass
class ExtractedWorkbook:
    tables: List[ExtractedTable]
    file_format: str
    sheet_count: int


def _cell_str(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _table_from_matrix(sheet_name: str, matrix: Sequence[Sequence[object]]) -> ExtractedTable | None:
    if not matrix:
        return None
    headers = [_cell_str(c) for c in matrix[0]]
    if not any(headers):
        return None
    rows: List[List[str]] = []
    for raw_row in matrix[1:]:
        row = [_cell_str(c) for c in raw_row]
        if any(row):
            rows.append(row)
    if not rows:
        return None
    col_count = max(len(headers), max((len(r) for r in rows), default=0))
    return ExtractedTable(
        sheet_name=sheet_name,
        headers=headers,
        rows=rows,
        row_start=2,
        row_end=1 + len(rows),
        col_count=col_count,
    )


def _extract_xlsx(path: str) -> ExtractedWorkbook:
    from openpyxl import load_workbook  # noqa: PLC0415

    wb = load_workbook(path, read_only=True, data_only=True)
    tables: List[ExtractedTable] = []
    try:
        for sheet in wb.worksheets:
            matrix: List[List[object]] = []
            for row in sheet.iter_rows(values_only=True):
                if any(cell is not None and str(cell).strip() for cell in row):
                    matrix.append(list(row))
            table = _table_from_matrix(sheet.title, matrix)
            if table:
                tables.append(table)
    finally:
        wb.close()
    return ExtractedWorkbook(tables=tables, file_format="xlsx", sheet_count=len(tables))


def _extract_csv(path: str) -> ExtractedWorkbook:
    tables: List[ExtractedTable] = []
    with open(path, newline="", encoding="utf-8-sig") as handle:
        reader = csv.reader(handle)
        matrix = [row for row in reader if any(cell.strip() for cell in row)]
    table = _table_from_matrix(Path(path).stem, matrix)
    if table:
        tables.append(table)
    return ExtractedWorkbook(tables=tables, file_format="csv", sheet_count=len(tables))


def extract_workbook(path: str) -> ExtractedWorkbook:
    suffix = Path(path).suffix.lower()
    if suffix == ".csv":
        return _extract_csv(path)
    if suffix in {".xlsx", ".xlsm"}:
        return _extract_xlsx(path)
    raise ValueError(f"Unsupported spreadsheet format: {suffix}")
