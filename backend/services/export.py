"""
backend/services/export.py  (TẠO MỚI)

Export kết quả predict / batch sang CSV, JSON, Excel.

Dùng trong pipeline hoặc thêm endpoint GET /export/{format}:
    from .services.export import export_predictions
"""

from __future__ import annotations

import csv
import json
from datetime import datetime
from io import StringIO, BytesIO
from pathlib import Path
from typing import Any


# ── CSV ───────────────────────────────────────────────────────────────────────

def to_csv_string(predictions: list[dict[str, Any]]) -> str:
    """Chuyển danh sách predict results thành CSV string."""
    if not predictions:
        return ""

    output = StringIO()
    fields = [
        "company", "ticker", "score",
        "short_term", "near_term",
        "confidence", "status",
        "risks", "opportunities",
        "executive_summary",
    ]
    writer = csv.DictWriter(output, fieldnames=fields, extrasaction="ignore")
    writer.writeheader()

    for p in predictions:
        trend = p.get("trend") or {}
        writer.writerow({
            "company":           p.get("company", ""),
            "ticker":            p.get("ticker", ""),
            "score":             p.get("score", ""),
            "short_term":        trend.get("short_term", "") if isinstance(trend, dict) else "",
            "near_term":         trend.get("near_term", "") if isinstance(trend, dict) else "",
            "confidence":        p.get("confidence", ""),
            "status":            p.get("status", ""),
            "risks":             "|".join(p.get("risks", [])),
            "opportunities":     "|".join(p.get("opportunities", [])),
            "executive_summary": p.get("executive_summary", ""),
        })
    return output.getvalue()


def to_csv_bytes(predictions: list[dict[str, Any]]) -> bytes:
    return to_csv_string(predictions).encode("utf-8-sig")  # utf-8-sig cho Excel


# ── JSON ──────────────────────────────────────────────────────────────────────

def to_json_bytes(predictions: list[dict[str, Any]], indent: int = 2) -> bytes:
    return json.dumps(predictions, ensure_ascii=False, indent=indent, default=str).encode("utf-8")


# ── Excel ─────────────────────────────────────────────────────────────────────

def to_excel_bytes(predictions: list[dict[str, Any]]) -> bytes:
    """
    Export sang .xlsx với formatting cơ bản.
    Yêu cầu: openpyxl (đã có trong requirements.txt)
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Predictions"

    # Header style
    header_fill = PatternFill("solid", fgColor="4F46E5")
    header_font = Font(bold=True, color="FFFFFF")

    headers = [
        "Company", "Ticker", "Score", "Short Term", "Near Term",
        "Confidence", "Status", "Risks", "Opportunities", "Summary",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, p in enumerate(predictions, 2):
        trend = p.get("trend") or {}
        score = p.get("score")

        ws.cell(row=row_idx, column=1, value=p.get("company", ""))
        ws.cell(row=row_idx, column=2, value=p.get("ticker", ""))

        score_cell = ws.cell(row=row_idx, column=3, value=round(score, 3) if score else "")
        if score is not None:
            if score >= 0.65:
                score_cell.fill = PatternFill("solid", fgColor="D1FAE5")
            elif score < 0.40:
                score_cell.fill = PatternFill("solid", fgColor="FEE2E2")

        st = trend.get("short_term", "") if isinstance(trend, dict) else ""
        nt = trend.get("near_term",  "") if isinstance(trend, dict) else ""
        ws.cell(row=row_idx, column=4, value=st)
        ws.cell(row=row_idx, column=5, value=nt)
        ws.cell(row=row_idx, column=6, value=round(p.get("confidence", 0), 3))
        ws.cell(row=row_idx, column=7, value=p.get("status", ""))
        ws.cell(row=row_idx, column=8, value="|".join(p.get("risks", [])))
        ws.cell(row=row_idx, column=9, value="|".join(p.get("opportunities", [])))
        ws.cell(row=row_idx, column=10, value=p.get("executive_summary", ""))

    # Column widths
    widths = [20, 12, 8, 12, 12, 12, 16, 30, 30, 50]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Save to disk ──────────────────────────────────────────────────────────────

def export_predictions(
    predictions: list[dict[str, Any]],
    output_dir: str | Path,
    formats: list[str] | None = None,
) -> dict[str, str]:
    """
    Export predictions ra file(s).

    Args:
        predictions: Danh sách kết quả từ TrendEngine.analyze()
        output_dir:  Thư mục đầu ra
        formats:     ["csv", "json", "excel"] — default: tất cả

    Returns:
        Dict mapping format → đường dẫn file
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    formats = formats or ["csv", "json", "excel"]

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    paths: dict[str, str] = {}

    if "csv" in formats:
        path = output_dir / f"predictions_{ts}.csv"
        path.write_bytes(to_csv_bytes(predictions))
        paths["csv"] = str(path)

    if "json" in formats:
        path = output_dir / f"predictions_{ts}.json"
        path.write_bytes(to_json_bytes(predictions))
        paths["json"] = str(path)

    if "excel" in formats:
        path = output_dir / f"predictions_{ts}.xlsx"
        path.write_bytes(to_excel_bytes(predictions))
        paths["excel"] = str(path)

    return paths