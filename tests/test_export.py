"""
tests/test_export.py
Tests cho export service (CSV, JSON, Excel).
"""

from __future__ import annotations

import csv
import json
import pytest
from io import StringIO
from pathlib import Path


SAMPLE = [
    {
        "company": "Vinamilk", "ticker": "VNM.HM",
        "score": 0.724, "confidence": 0.87, "status": "OK",
        "trend": {"short_term": "bullish", "near_term": "bullish"},
        "risks": ["cost_up"], "opportunities": ["revenue_up", "eps_beat"],
        "executive_summary": "Tăng trưởng tốt.",
    },
    {
        "company": "Acme Corp", "ticker": "ACME.US",
        "score": 0.32, "confidence": 0.70, "status": "OK",
        "trend": {"short_term": "bearish", "near_term": "bearish"},
        "risks": ["revenue_down", "eps_miss"], "opportunities": [],
        "executive_summary": "Suy giảm.",
    },
    {
        "company": "Unknown Co", "ticker": None,
        "score": None, "confidence": 0.0, "status": "INSUFFICIENT_DATA",
        "trend": None, "risks": [], "opportunities": [],
        "executive_summary": None,
    },
]


# ── CSV ───────────────────────────────────────────────────────────────────────

class TestToCSV:

    def test_returns_string(self):
        from backend.services.export import to_csv_string
        result = to_csv_string(SAMPLE)
        assert isinstance(result, str)
        assert len(result) > 0

    def test_has_header_row(self):
        from backend.services.export import to_csv_string
        csv_str = to_csv_string(SAMPLE)
        reader = csv.DictReader(StringIO(csv_str))
        assert "company"    in reader.fieldnames
        assert "score"      in reader.fieldnames
        assert "short_term" in reader.fieldnames
        assert "confidence" in reader.fieldnames

    def test_row_count_matches(self):
        from backend.services.export import to_csv_string
        csv_str = to_csv_string(SAMPLE)
        rows = list(csv.DictReader(StringIO(csv_str)))
        assert len(rows) == len(SAMPLE)

    def test_company_names_correct(self):
        from backend.services.export import to_csv_string
        rows = list(csv.DictReader(StringIO(to_csv_string(SAMPLE))))
        companies = [r["company"] for r in rows]
        assert "Vinamilk" in companies
        assert "Acme Corp" in companies

    def test_risks_joined_with_pipe(self):
        from backend.services.export import to_csv_string
        rows = list(csv.DictReader(StringIO(to_csv_string(SAMPLE))))
        acme = next(r for r in rows if r["company"] == "Acme Corp")
        assert "revenue_down" in acme["risks"]
        assert "|" in acme["risks"]

    def test_empty_list_returns_empty_string(self):
        from backend.services.export import to_csv_string
        assert to_csv_string([]) == ""

    def test_bytes_has_utf8_bom(self):
        from backend.services.export import to_csv_bytes
        b = to_csv_bytes(SAMPLE)
        # utf-8-sig BOM = \xef\xbb\xbf
        assert b[:3] == b"\xef\xbb\xbf"

    def test_none_score_handled(self):
        from backend.services.export import to_csv_string
        data = [{"company": "X", "score": None, "status": "INSUFFICIENT_DATA",
                 "trend": None, "risks": [], "opportunities": [],
                 "confidence": 0.0, "executive_summary": ""}]
        csv_str = to_csv_string(data)
        assert "X" in csv_str

    def test_none_trend_handled(self):
        from backend.services.export import to_csv_string
        result = to_csv_string(SAMPLE)
        # Không được raise exception dù trend=None
        assert "Unknown Co" in result


# ── JSON ──────────────────────────────────────────────────────────────────────

class TestToJSON:

    def test_returns_valid_json(self):
        from backend.services.export import to_json_bytes
        b = to_json_bytes(SAMPLE)
        parsed = json.loads(b.decode("utf-8"))
        assert isinstance(parsed, list)
        assert len(parsed) == len(SAMPLE)

    def test_preserves_all_fields(self):
        from backend.services.export import to_json_bytes
        parsed = json.loads(to_json_bytes(SAMPLE).decode())
        first = parsed[0]
        assert first["company"]  == "Vinamilk"
        assert first["score"]    == 0.724
        assert first["status"]   == "OK"

    def test_empty_list(self):
        from backend.services.export import to_json_bytes
        parsed = json.loads(to_json_bytes([]).decode())
        assert parsed == []

    def test_none_values_serialized(self):
        from backend.services.export import to_json_bytes
        data = [{"company": "X", "score": None, "trend": None}]
        parsed = json.loads(to_json_bytes(data).decode())
        assert parsed[0]["score"] is None


# ── Excel ─────────────────────────────────────────────────────────────────────

class TestToExcel:

    def test_returns_bytes(self):
        from backend.services.export import to_excel_bytes
        b = to_excel_bytes(SAMPLE)
        assert isinstance(b, bytes)
        assert len(b) > 0

    def test_valid_xlsx_magic_bytes(self):
        from backend.services.export import to_excel_bytes
        b = to_excel_bytes(SAMPLE)
        # .xlsx là ZIP file — magic bytes PK\x03\x04
        assert b[:4] == b"PK\x03\x04"

    def test_readable_by_openpyxl(self):
        import openpyxl
        from io import BytesIO
        from backend.services.export import to_excel_bytes
        b = to_excel_bytes(SAMPLE)
        wb = openpyxl.load_workbook(BytesIO(b))
        ws = wb.active
        # Row 1 = headers, row 2+ = data
        assert ws.max_row >= len(SAMPLE) + 1
        assert ws.max_column >= 8

    def test_header_row_content(self):
        import openpyxl
        from io import BytesIO
        from backend.services.export import to_excel_bytes
        wb = openpyxl.load_workbook(BytesIO(to_excel_bytes(SAMPLE)))
        ws = wb.active
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert "Company" in headers
        assert "Score"   in headers
        assert "Status"  in headers

    def test_data_in_rows(self):
        import openpyxl
        from io import BytesIO
        from backend.services.export import to_excel_bytes
        wb = openpyxl.load_workbook(BytesIO(to_excel_bytes(SAMPLE)))
        ws = wb.active
        companies = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
        assert "Vinamilk" in companies

    def test_empty_list_creates_headers_only(self):
        import openpyxl
        from io import BytesIO
        from backend.services.export import to_excel_bytes
        wb = openpyxl.load_workbook(BytesIO(to_excel_bytes([])))
        ws = wb.active
        assert ws.max_row == 1  # only header


# ── export_predictions (file output) ─────────────────────────────────────────

class TestExportPredictions:

    def test_creates_csv_file(self, tmp_path):
        from backend.services.export import export_predictions
        paths = export_predictions(SAMPLE, tmp_path, formats=["csv"])
        assert "csv" in paths
        assert Path(paths["csv"]).exists()
        assert Path(paths["csv"]).stat().st_size > 0

    def test_creates_json_file(self, tmp_path):
        from backend.services.export import export_predictions
        paths = export_predictions(SAMPLE, tmp_path, formats=["json"])
        assert "json" in paths
        assert Path(paths["json"]).exists()

    def test_creates_excel_file(self, tmp_path):
        from backend.services.export import export_predictions
        paths = export_predictions(SAMPLE, tmp_path, formats=["excel"])
        assert "excel" in paths
        assert Path(paths["excel"]).suffix == ".xlsx"

    def test_creates_all_formats_by_default(self, tmp_path):
        from backend.services.export import export_predictions
        paths = export_predictions(SAMPLE, tmp_path)
        assert set(paths.keys()) == {"csv", "json", "excel"}
        for p in paths.values():
            assert Path(p).exists()

    def test_creates_output_dir_if_missing(self, tmp_path):
        from backend.services.export import export_predictions
        new_dir = tmp_path / "nested" / "output"
        assert not new_dir.exists()
        export_predictions(SAMPLE, new_dir, formats=["csv"])
        assert new_dir.exists()

    def test_filename_contains_timestamp(self, tmp_path):
        from backend.services.export import export_predictions
        paths = export_predictions(SAMPLE, tmp_path, formats=["csv"])
        filename = Path(paths["csv"]).name
        assert "predictions_" in filename
        # Format: predictions_YYYYMMDD_HHMMSS.csv
        assert len(filename) > len("predictions_.csv")

    def test_json_content_correct(self, tmp_path):
        from backend.services.export import export_predictions
        paths = export_predictions(SAMPLE, tmp_path, formats=["json"])
        content = json.loads(Path(paths["json"]).read_bytes())
        assert len(content) == len(SAMPLE)
        assert content[0]["company"] == "Vinamilk"